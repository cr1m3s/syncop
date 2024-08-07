#!/usr/bin/python3

import sys
import openpyxl
from os.path import exists

#
#   takes LO and creates .xlsx file with datacollection
#   no formatting support and node names
#   file /data/sfng/log/{LO}/datacollection{LO}-to-excel.xlsx should exist
#   .xlsx file creates in /data/sfng/{LO}/ 
#

if __name__ == '__main__':
    
    filename = ''
    LO = ''
    SPP_VERSION = ''
    datacollection = []
    ulocData = {}

    if len(sys.argv) != 2:
        print("Usage:\n./excel_format.py {LO}")
        sys.exit(-1)
    else:
        LO = sys.argv[1]

    #GET DATACOLLECTION
    filename = f"/data/sfng/log/{LO}/datacollection_{LO}-to-excel.txt"
    if not exists(filename):
        print(f"No such file: {filename}\n")
        exit(-1)

    with open(filename) as f:
        for line in f:
            datacollection.append(line.rstrip())
    
    #GET LEGACY ORDER
    lo = ''
    for item in datacollection:
        if "Sales Order" in item:
            lo = item.split(",")[1]
            break

    # GET SPP version FROM LOGS FILE
    with open(f"/data/log/NokiaNSW/Ansible/ansible_{LO}.log") as f:
        spp_raw = f.read().split("\n")[0]
        SPP_VERSION = spp_raw.split(" ")[-2]
        print(SPP_VERSION)

    #MAKE DICTIONARY dict["U{num}"][VALUE]
    for i in range(len(datacollection)):
        if ":U" in datacollection[i]:
            ulocData[datacollection[i][1:]] = datacollection[i+1].split("\t")
            i+=1
    
    flag = True
    
    # Default template for Gen 10
    template = "Nokia_DataCollection_Cust_Cascadelake_.xlsx"
    
    for key in ulocData.keys():
        print(f"{key}: {ulocData[key]}")
        if flag:
            flag = False
            # Pick another template based on the server Generation
            for i in ulocData[key]:
                if "Gen11" in i:
                    template = "Nokia_DataCollection_Cust_Sapphire_Rapids_.xlsx" 
                    break
                elif "Gen10 Plus" in i:
                    template = "Nokia_DataCollection_Cust_Icalake_.xlsx"
                    break
    print(f"Using template: {template}") 

    writeOrder = list(ulocData.keys())
    writeOrder.sort()
    ulocCell = {}
    
    #OPEN TEMPLATE .xlsx FILE
    workbook = openpyxl.load_workbook(template)
    sheet = workbook.active

    #START POINT TO WRITE DATA
    for i in sheet.iter_cols():
        for j in i:
            if j.value in writeOrder :
                ulocCell[j.value] = [j.column, j.row]

    #SKIP FIRST TWO CELLS
    for key in ulocCell.keys():
        sheet.cell(row = ulocCell[key][1], column = ulocCell[key][0]+3).value = int(lo)

    for i in writeOrder:
        startIter = 5
        ccolumn = ulocCell[i][0]
        crow = ulocCell[i][1]           
        for data in ulocData[i]:    
            sheet.cell(row = crow, column = ccolumn + startIter).value = data
            startIter += 1
        sheet.cell(row=crow, column=ccolumn + 20).value = SPP_VERSION


    
    result = f"/data/sfng/log/{LO}/{template.split('.')[0]}{LO}.xlsx"
    workbook.save(result)
    print(f"Datacollection stored in {result}\n")    
 
    workbook.close()
